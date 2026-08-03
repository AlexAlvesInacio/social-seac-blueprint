#!/usr/bin/env python3
"""Importa o cadastro e o histórico de entregas da planilha legada da SEAC.

Contexto e decisões: docs/13_IMPORTACAO_PLANILHA_LEGADA.md.

Uso:
    .venv/bin/python scripts/importar-planilha.py "CESTAS SEAC 2026.xlsx" [--aplicar]

Sem --aplicar faz um ensaio: percorre tudo, grava o relatório e desfaz. Com
--aplicar, confirma. O ensaio é o modo padrão de propósito.

Exige as variáveis SEAC_DB_URL (conexão) e SEAC_ADMIN_ID (uuid do administrador
que assina a carga na auditoria).
"""

from __future__ import annotations

import argparse
import collections
import datetime
import os
import re
import sys

import openpyxl
import psycopg

# A tabela de retiradas começa abaixo do texto de regras da aba.
LINHA_CABECALHO_CONTROLE = 14

# Nome usado quando a pessoa é atendida antes de ter cadastro.
PLACEHOLDER_NOME = "PRIMEIRA VEZ"

# Documentos que são preenchimento, não identificação.
DOCUMENTOS_INVALIDOS = {"10", "20", "30", "40", "50", "60", "XXX", "XXXX", "YYY"}

BENEFICIO_POR_TIPO = {"comum": "Cesta Padrão", "extra": "Cesta Extra"}
TIPO_CADASTRO_POR_TIPO = {"comum": "definitivo", "extra": "extra"}


def so_alnum(valor) -> str:
    if valor in (None, ""):
        return ""
    return re.sub(r"[^0-9A-Za-z]", "", str(valor)).upper()


def cpf_valido(numero: str) -> bool:
    """Dígitos verificadores. Mesmo cálculo de private.cpf_valido no banco."""
    if len(numero) != 11 or not numero.isdigit() or len(set(numero)) == 1:
        return False
    for corte in (9, 10):
        soma = sum(int(numero[i]) * ((corte + 1) - i) for i in range(corte))
        if (soma * 10) % 11 % 10 != int(numero[corte]):
            return False
    return True


def indices(cabecalho) -> dict[str, int]:
    return {
        str(v).strip().upper().replace("\n", " "): i
        for i, v in enumerate(cabecalho)
        if v
    }


def ler_nomes_conhecidos(wb) -> dict[str, str]:
    """Nome por RG, varrendo as abas de cadastro.

    Serve para recuperar o nome de quem foi atendido como "PRIMEIRA VEZ" e
    depois cadastrado — 31 das 41 ocorrências se resolvem assim.
    """
    nomes: dict[str, str] = {}
    for aba in (
        "BANCO DE DADOS",
        "Cópia de BANCO DE DADOS",
        "BKP BANCO DE DADOS 20241006",
        "ADICIONADOS",
    ):
        if aba not in wb.sheetnames:
            continue
        it = wb[aba].iter_rows(values_only=True)
        idx = indices(next(it))
        i_rg, i_nome = idx.get("RG"), idx.get("NOME")
        if i_rg is None or i_nome is None:
            continue
        for linha in it:
            rg, nome = so_alnum(linha[i_rg]), str(linha[i_nome] or "").strip()
            if rg and nome and PLACEHOLDER_NOME not in nome.upper():
                nomes.setdefault(rg, nome)
    return nomes


def ler_cpfs_conhecidos(wb) -> dict[str, str]:
    """CPF por RG, para enriquecer o cadastro importado."""
    cpfs: dict[str, str] = {}
    for aba in ("BANCO DE DADOS", "Cópia de BANCO DE DADOS"):
        if aba not in wb.sheetnames:
            continue
        it = wb[aba].iter_rows(values_only=True)
        idx = indices(next(it))
        i_rg, i_cpf = idx.get("RG"), idx.get("CPF")
        if i_rg is None or i_cpf is None:
            continue
        for linha in it:
            rg, cpf = so_alnum(linha[i_rg]), so_alnum(linha[i_cpf])
            if rg and cpf_valido(cpf):
                cpfs.setdefault(rg, cpf)
    return cpfs


class Pessoa:
    __slots__ = ("chave", "nome", "documento", "tipo_documento", "cpf", "retiradas")

    def __init__(self, chave: str):
        self.chave = chave
        self.nome = ""
        self.documento = ""
        self.tipo_documento = "rg"
        self.cpf: str | None = None
        self.retiradas: list[tuple[datetime.datetime, str, str]] = []

    @property
    def tipo_cadastro(self) -> str:
        """Definido pela ÚLTIMA retirada.

        58 das 61 pessoas que receberam os dois tipos seguem a ordem
        extra → comum, que é o fluxo avaliação → definitivo.
        """
        _, tipo, _ = max(self.retiradas)
        return TIPO_CADASTRO_POR_TIPO[tipo]


def ler_planilha(caminho: str):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    nomes = ler_nomes_conhecidos(wb)
    cpfs = ler_cpfs_conhecidos(wb)

    linhas = list(wb["CONTROLE"].iter_rows(min_row=LINHA_CABECALHO_CONTROLE, values_only=True))
    idx = indices(linhas[0])
    i_rg, i_nome, i_data = idx["RG"], idx["NOME"], idx["DATA RETIRADA"]
    i_unico = idx["UNICO"]
    i_dif = next(i for chave, i in idx.items() if "DIFERENCIADA" in chave)

    pessoas: dict[str, Pessoa] = {}
    rejeitados: list[tuple[str, str, str]] = []
    # A planilha tem chaves UNICO repetidas. A RPC é idempotente e absorve a
    # repetição, mas é melhor a equipe saber: pode ser digitação duplicada ou
    # duas retiradas reais no mesmo dia registradas com a mesma sequência.
    vistas: set[str] = set()
    repetidas: list[tuple[str, str]] = []

    for linha in linhas[1:]:
        nome_bruto = str(linha[i_nome] or "").strip()
        rg = so_alnum(linha[i_rg])
        if not nome_bruto and not rg:
            continue

        chave_externa = str(linha[i_unico] or "").strip()
        rotulo = f"{nome_bruto or '(sem nome)'} / RG {rg or '(vazio)'}"

        if not rg or rg in DOCUMENTOS_INVALIDOS:
            rejeitados.append((rotulo, chave_externa, "documento ausente ou de preenchimento"))
            continue
        if not isinstance(linha[i_data], datetime.datetime):
            rejeitados.append((rotulo, chave_externa, "data de retirada ausente ou inválida"))
            continue
        if not chave_externa:
            rejeitados.append((rotulo, chave_externa, "campo UNICO vazio: sem chave, a reimportação duplicaria"))
            continue

        nome = nome_bruto
        if not nome or PLACEHOLDER_NOME in nome.upper():
            nome = nomes.get(rg, "")
            if not nome:
                rejeitados.append((rotulo, chave_externa, "nome de preenchimento e não encontrado em outra aba"))
                continue

        pessoa = pessoas.get(rg)
        if pessoa is None:
            pessoa = pessoas[rg] = Pessoa(rg)
            pessoa.nome = nome
            # 42 RGs de 11 dígitos passam no dígito verificador: são CPF
            # digitado no campo errado. Entram como CPF, que é o documento
            # para onde a SEAC está migrando.
            if cpf_valido(rg):
                pessoa.documento, pessoa.tipo_documento = rg, "cpf"
            else:
                pessoa.documento, pessoa.tipo_documento = rg, "rg"
                pessoa.cpf = cpfs.get(rg)
        elif not pessoa.nome:
            pessoa.nome = nome

        if chave_externa in vistas:
            repetidas.append((rotulo, chave_externa))
        vistas.add(chave_externa)

        tipo = "extra" if linha[i_dif] not in (None, "") else "comum"
        pessoa.retiradas.append((linha[i_data], tipo, chave_externa))

    return pessoas, rejeitados, repetidas


def importar(conn, pessoas: dict[str, Pessoa], admin_id: str):
    contagem = collections.Counter()
    falhas: list[tuple[str, str]] = []
    avisos: list[tuple[str, str]] = []

    with conn.cursor() as cur:
        cur.execute(
            "select set_config('request.jwt.claims', %s, true)",
            ('{"sub":"%s","role":"authenticated"}' % admin_id,),
        )
        cur.execute("set local role authenticated")

        for pessoa in pessoas.values():
            try:
                with conn.transaction():
                    cur.execute(
                        """
                        select familia_id, assistido_id
                        from public.criar_familia_com_responsavel(
                          %s, %s, %s::public.pessoa_tipo_documento, %s,
                          %s::public.assistido_tipo_cadastro,
                          null, null, null, null, null, null, null, null
                        )
                        """,
                        (pessoa.nome, pessoa.nome, pessoa.tipo_documento,
                         pessoa.documento, pessoa.tipo_cadastro),
                    )
                    _, assistido_id = cur.fetchone()
                    contagem["pessoas"] += 1

                    if pessoa.cpf:
                        # Em savepoint próprio: dois RGs da planilha apontam
                        # para o mesmo CPF, e a colisão não pode derrubar a
                        # pessoa nem o histórico de retiradas dela.
                        try:
                            with conn.transaction():
                                cur.execute(
                                    "update public.pessoas set cpf = %s where documento_normalizado = %s",
                                    (pessoa.cpf, so_alnum(pessoa.documento)),
                                )
                            contagem["cpf_enriquecido"] += 1
                        except Exception:  # noqa: BLE001
                            contagem["cpf_em_conflito"] += 1
                            avisos.append(
                                (f"{pessoa.nome} / {pessoa.documento}",
                                 f"CPF {pessoa.cpf} já pertence a outra pessoa — importada sem CPF")
                            )

                    for data, tipo, chave in sorted(pessoa.retiradas):
                        cur.execute(
                            "select public.carregar_entrega_historica(%s, %s, %s, %s, %s)",
                            (assistido_id, BENEFICIO_POR_TIPO[tipo], data, chave,
                             "Importado da planilha legada"),
                        )
                        contagem["entregas"] += 1
            except Exception as erro:  # noqa: BLE001 — a falha vira relatório
                falhas.append((f"{pessoa.nome} / {pessoa.documento}", str(erro).strip().splitlines()[0]))

    # Confere o resultado ainda dentro da transação: no ensaio é a única
    # chance de ver o que a importação produziria.
    with conn.cursor() as cur:
        cur.execute("set local role postgres")
        cur.execute("""
            select
              (select count(*) from public.pessoas),
              (select count(*) from public.familias),
              (select count(*) from public.entregas),
              (select count(*) from public.assistidos where tipo_cadastro='definitivo'),
              (select count(*) from public.assistidos where tipo_cadastro='extra'),
              (select count(*) from (
                 select pessoa_id from public.entregas
                 group by pessoa_id
                 having max(criado_em) > now() - interval '25 days') b),
              (select coalesce(sum(saldo),0) from public.beneficios),
              (select count(*) from public.movimentacoes_estoque)
        """)
        conferencia = cur.fetchone()

    return contagem, falhas, avisos, conferencia


def main() -> int:
    ap = argparse.ArgumentParser(description="Importa a planilha legada da SEAC.")
    ap.add_argument("planilha")
    ap.add_argument("--aplicar", action="store_true",
                    help="confirma a transação; sem isto, faz ensaio e desfaz")
    ap.add_argument("--relatorio", default="relatorio-importacao.txt")
    args = ap.parse_args()

    url, admin = os.environ.get("SEAC_DB_URL"), os.environ.get("SEAC_ADMIN_ID")
    if not url or not admin:
        print("Defina SEAC_DB_URL e SEAC_ADMIN_ID.", file=sys.stderr)
        return 2

    pessoas, rejeitados, repetidas = ler_planilha(args.planilha)
    total_retiradas = sum(len(p.retiradas) for p in pessoas.values())
    print(f"planilha: {len(pessoas)} pessoas, {total_retiradas} retiradas, "
          f"{len(rejeitados)} linhas rejeitadas, {len(repetidas)} chaves repetidas")

    with psycopg.connect(url) as conn:
        contagem, falhas, avisos, conf = importar(conn, pessoas, admin)
        if args.aplicar:
            conn.commit()
            print("ALTERAÇÕES CONFIRMADAS")
        else:
            conn.rollback()
            print("ENSAIO — nada foi gravado (use --aplicar para confirmar)")

    with open(args.relatorio, "w", encoding="utf-8") as saida:
        saida.write(f"Importação da planilha legada — {datetime.date.today()}\n")
        saida.write(f"{'APLICADA' if args.aplicar else 'ENSAIO'}\n\n")
        saida.write(f"pessoas criadas .... {contagem['pessoas']}\n")
        saida.write(f"CPF enriquecido .... {contagem['cpf_enriquecido']}\n")
        saida.write(f"entregas carregadas  {contagem['entregas']}\n")
        saida.write(f"CPF em conflito .... {contagem['cpf_em_conflito']}\n\n")
        saida.write(f"REJEITADOS NA LEITURA ({len(rejeitados)})\n")
        for rotulo, chave, motivo in rejeitados:
            saida.write(f"  {rotulo} [{chave}]: {motivo}\n")
        saida.write(f"\nCHAVES UNICO REPETIDAS ({len(repetidas)}) — importadas uma vez só\n")
        for rotulo, chave in repetidas:
            saida.write(f"  {rotulo} [{chave}]\n")
        saida.write(f"\nAVISOS ({len(avisos)}) — importados, mas com ressalva\n")
        for rotulo, aviso in avisos:
            saida.write(f"  {rotulo}: {aviso}\n")
        saida.write(f"\nFALHAS NA GRAVAÇÃO ({len(falhas)})\n")
        for rotulo, erro in falhas:
            saida.write(f"  {rotulo}: {erro}\n")

    rotulos = ("pessoas", "famílias", "entregas", "definitivos", "extras",
               "bloqueados por prazo", "saldo total de benefícios",
               "movimentações de estoque")
    print("\nestado do banco ao fim da importação:")
    for rotulo, valor in zip(rotulos, conf):
        print(f"  {rotulo:.<30} {valor}")
    print()
    print(f"pessoas={contagem['pessoas']} entregas={contagem['entregas']} "
          f"cpf={contagem['cpf_enriquecido']} rejeitados={len(rejeitados)} falhas={len(falhas)}")
    print(f"relatório em {args.relatorio}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
